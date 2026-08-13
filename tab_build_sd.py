"""
Skrip admin (sekali jalan): isi Google Sheet TABUNGAN SD dari file Excel master.
Prasyarat: spreadsheet SD (tab_config.SD_SPREADSHEET_ID) sudah di-share ke
service account (Editor).

Jalankan dari folder r5401-report-to-excel:
    python3 tab_build_sd.py            # isi tab SD 1..6 2026
    python3 tab_build_sd.py 2027       # buat T.A. berikutnya (saldo awal = saldo Juni)
"""
import os
import sys

import openpyxl

import tab_config as C
import tab_sheet as SL

JENJANG = "SD"
SID = C.SD_SPREADSHEET_ID
SRC = os.path.join(os.path.dirname(__file__), "..",
                   "TABUNGAN SD INSAN AMANAH UNTUK MAS OKI.xlsx")
# kelas -> nama tab di file Excel master
SRC_TABS = {1: "KELAS 1", 2: "KELAS 2", 3: "KELAS 3",
            4: "KELAS 4", 5: "KELAS 5", 6: "KELAS 6"}


def _data_start(ws):
    for r in range(1, 9):
        v = ws.cell(r, 2).value
        if isinstance(v, (int, float)) or (isinstance(v, str) and str(v).strip().isdigit()):
            return r
    return 3


def read_roster(path, kelas):
    ws = openpyxl.load_workbook(path, data_only=True)[SRC_TABS[kelas]]
    start = _data_start(ws)
    has_saldo = kelas != 1          # KELAS 1 = siswa baru, tanpa saldo awal
    rows = []
    for r in range(start, ws.max_row + 1):
        b, c = ws.cell(r, 2).value, ws.cell(r, 3).value
        if b in (None, "") and c in (None, ""):
            continue
        if isinstance(b, float) and b == int(b):
            induk = str(int(b))
        else:
            induk = str(b).strip() if b is not None else ""
        nama = str(c).strip() if c is not None else ""
        saldo = 0
        if has_saldo:
            d = ws.cell(r, 4).value
            saldo = d if isinstance(d, (int, float)) else 0
        rows.append((induk, nama, saldo))
    return rows


def main():
    year = int(sys.argv[1]) if len(sys.argv) > 1 else C.START_YEAR
    book = SL.open_book(SID)
    print("Terhubung:", book.title)

    for kelas in C.SD_KELAS_LIST:
        if year == C.START_YEAR:
            roster = read_roster(SRC, kelas)
        else:
            prev = book.worksheet(C.tab_name(kelas, year - 1, JENJANG))
            roster = SL.last_saldo_of_year(prev, year - 1)
        SL.build_tab(book, kelas, year, roster, JENJANG)
        print(f"  ✓ {C.tab_name(kelas, year, JENJANG)}: {len(roster)} siswa")

    if year == C.START_YEAR:
        for junk in ("Sheet1", "Sheet"):
            try:
                book.del_worksheet(book.worksheet(junk))
            except Exception:
                pass
    print("Selesai:", f"https://docs.google.com/spreadsheets/d/{SID}")


if __name__ == "__main__":
    main()
