"""
Skrip admin (sekali jalan): isi Google Sheet TABUNGAN dari file Excel asli.
Prasyarat: spreadsheet sudah di-share ke service account (Editor).

Jalankan dari folder r5401-report-to-excel:
    python3 tab_build.py            # isi tab SMP 7/8/9 2026 dari Excel
    python3 tab_build.py 2027       # buat T.A. berikutnya (saldo awal = saldo Juni)
"""
import os
import sys

import openpyxl

import tab_config as C
import tab_sheet as SL

# file Excel asli ada satu level di atas folder ini
SRC = os.path.join(os.path.dirname(__file__), "..", "TABUNGAN SMP INSAN AMANAH.xlsx")


def read_roster(path, kelas):
    ws = openpyxl.load_workbook(path, data_only=True)[str(kelas)]
    rows = []
    for r in range(4, ws.max_row + 1):
        induk, nama = ws.cell(r, 2).value, ws.cell(r, 3).value
        if (induk in (None, "")) and (nama in (None, "")):
            continue
        rows.append((str(induk).strip() if induk is not None else "",
                     str(nama).strip() if nama is not None else "",
                     ws.cell(r, 4).value or 0))
    return rows


def main():
    year = int(sys.argv[1]) if len(sys.argv) > 1 else C.START_YEAR
    book = SL.open_book()
    print("Terhubung:", book.title)

    for kelas in C.KELAS_LIST:
        if year == C.START_YEAR:
            roster = read_roster(SRC, kelas)
        else:
            prev = book.worksheet(C.tab_name(kelas, year - 1))
            roster = SL.last_saldo_of_year(prev, year - 1)   # saldo Juni -> saldo awal
        SL.build_tab(book, kelas, year, roster)
        print(f"  ✓ {C.tab_name(kelas, year)}: {len(roster)} siswa")

    if year == C.START_YEAR:
        for junk in ("Sheet1", "Sheet"):
            try:
                book.del_worksheet(book.worksheet(junk))
            except Exception:
                pass
    print("Selesai:", f"https://docs.google.com/spreadsheets/d/{C.SPREADSHEET_ID}")


if __name__ == "__main__":
    main()
