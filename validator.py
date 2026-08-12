"""Rekonsiliasi: master siswa (xlsx) di-join ke laporan harian R-5401 (txt).

Join key: No. Pelanggan (laporan) == NO VA (master siswa). Kolom komponen
output (BPP/Kegiatan/Tabungan) mengikuti langsung nama kolom di master siswa,
tanpa ditukar.

Tanpa pandas/numpy — baca xlsx langsung dengan openpyxl (read_only) supaya
tetap ringan, konsisten dengan gaya parser.py.
"""
import re
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_KELAS = re.compile(r"\b(I|II|III|IV|V|VI)\s+([A-F])\b")


def _kelas(nama):
    m = _KELAS.search(str(nama).upper())
    return f"{m.group(1)} {m.group(2)}" if m else ""


def _strip_kelas(nama):
    return re.sub(r"\s*\b(I|II|III|IV|V|VI)\s+[A-F]\b.*$", "", str(nama).strip()).strip()


def _to_int(v):
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return 0


def parse_master_siswa(file_bytes):
    """Bytes xlsx (skema NO VA | NAMA | BPP | KEGIATAN | TABUNGAN) -> dict {no_va: row}."""
    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        if not row or row[0] is None or row[1] is None:
            continue
        try:
            no_va = int(row[0])
        except (TypeError, ValueError):
            continue
        nama = str(row[1]).strip()
        bpp = _to_int(row[2]) if len(row) > 2 else 0
        kegiatan = _to_int(row[3]) if len(row) > 3 else 0
        tabungan = _to_int(row[4]) if len(row) > 4 else 0
        out[no_va] = {
            "no_va": no_va, "nama": nama, "nama_bersih": _strip_kelas(nama),
            "kelas": _kelas(nama), "BPP": bpp, "KEGIATAN": kegiatan, "TABUNGAN": tabungan,
        }
    wb.close()
    return out


STATUS_SESUAI = "Sesuai"
STATUS_KURANG = "Kurang"
STATUS_LEBIH = "Lebih"


def _resolve_no_va(cust, kode, level):
    """Tentukan NO VA (kunci master) dari 'No. Pelanggan' laporan.

    - level 'sd' : No. Pelanggan pada laporan SUDAH berupa NO VA penuh -> dipakai langsung.
    - level 'smp': laporan hanya memuat kode pelanggan pendek (mis. '0318'); NO VA penuh
      dibentuk dari kode sekolah + kode pelanggan 4 digit (mis. 63713 + 0318 = 637130318).
    """
    c = str(cust or "").strip()
    if not c.isdigit():
        return None
    if level == "smp":
        k = str(kode or "").strip()
        if not k.isdigit():
            return None
        # Bila cust sudah berupa NO VA penuh (diawali kode sekolah, mis. dari laporan
        # BCA VA '63713-0388' -> '637130388'), pakai langsung; jangan diprefix lagi.
        if len(c) > len(k) and c.startswith(k):
            return int(c)
        return int(k + c.zfill(4))
    return int(c)


def reconcile_pembayaran(report_rows, master, kode=None, level="sd"):
    """Join baris laporan R-5401 (dari parser.parse_report) ke master siswa.

    report_rows: [no, no_pelanggan, nama, nilai, tgl(date), waktu(time), jam(int), lokasi, ket1, ket2]
    kode : kode sekolah dari meta laporan (dipakai saat level='smp' untuk membentuk NO VA penuh).
    level: 'sd' (No. Pelanggan = NO VA langsung) atau 'smp' (NO VA = kode + No. Pelanggan).
    -> list of dict siap-lapor (BPP/Kegiatan/Tabungan/Total Tagihan/Nilai Bayar/Selisih/Status).
    """
    out = []
    for row in report_rows:
        no, cust, nama_rpt, nilai, tgl, waktu, jam, lok, k1, k2 = row
        no_va = _resolve_no_va(cust, kode, level)
        m = master.get(no_va) if no_va is not None else None
        bpp = m["BPP"] if m else 0
        kegiatan = m["KEGIATAN"] if m else 0
        tabungan = m["TABUNGAN"] if m else 0
        total_tagihan = bpp + kegiatan + tabungan
        nilai_bayar = int(round(nilai))
        selisih = nilai_bayar - total_tagihan
        status = STATUS_SESUAI if selisih == 0 else (STATUS_LEBIH if selisih > 0 else STATUS_KURANG)
        out.append({
            "no_pelanggan": cust, "nama": m["nama_bersih"] if m else nama_rpt,
            "tgl": tgl, "waktu": waktu, "lokasi": lok,
            "bpp": bpp, "kegiatan": kegiatan, "tabungan": tabungan,
            "total_tagihan": total_tagihan, "nilai_bayar": nilai_bayar, "selisih": selisih,
            "status": status, "matched": m is not None, "ket1": k1, "ket2": k2,
        })
    return out


# ---------- styling (selaras dengan parser.py) ----------
_A = "Arial"
_HFILL = PatternFill("solid", fgColor="1F4E5F")
_HFONT = Font(name=_A, bold=True, color="FFFFFF", size=10)
_TITLE = Font(name=_A, bold=True, size=14, color="1F4E5F")
_SUB = Font(name=_A, italic=True, size=9, color="595959")
_BASE = Font(name=_A, size=10)
_BOLD = Font(name=_A, bold=True, size=10)
_TFILL = PatternFill("solid", fgColor="D9E7EC")
_OK = PatternFill("solid", fgColor="E8F5EC")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_C = Alignment("center", vertical="center")
_L = Alignment("left", vertical="center")
_R = Alignment("right", vertical="center")
_HEADERS = ["No", "No. Pelanggan", "Nama Pelanggan", "Tgl Transaksi", "Waktu", "Lokasi",
            "BPP", "Kegiatan", "Tabungan", "Total Tagihan", "Nilai Bayar", "Selisih",
            "Status", "Keterangan 1", "Keterangan 2"]
_MONEY_COLS = {7, 8, 9, 10, 11, 12}


def build_recon_workbook(rows, meta, only_sesuai=True):
    """rows: hasil reconcile_pembayaran(). Sheet Transaksi hanya berisi baris
    Status == Sesuai bila only_sesuai=True; sheet Ringkasan tetap merangkum semua baris."""
    shown = [r for r in rows if r["status"] == STATUS_SESUAI] if only_sesuai else rows

    wb = Workbook()
    ws = wb.active
    ws.title = "Transaksi"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "TRANSAKSI R-5401 vs MASTER SISWA"
    ws["A1"].font = _TITLE
    ws["A2"] = (f"{meta.get('nama_pt','') or ''} ({meta.get('kode','') or '—'})  •  "
                f"Tanggal: {meta.get('tanggal_label','') or '—'}"
                + ("  •  hanya status Sesuai" if only_sesuai else ""))
    ws["A2"].font = _SUB
    HROW = 4
    for c, h in enumerate(_HEADERS, 1):
        cell = ws.cell(HROW, c, h)
        cell.fill = _HFILL; cell.font = _HFONT; cell.alignment = _C; cell.border = _BORDER
    r0 = HROW + 1
    for i, row in enumerate(shown):
        r = r0 + i
        vals = [i + 1, row["no_pelanggan"], row["nama"], row["tgl"], row["waktu"], row["lokasi"],
                row["bpp"], row["kegiatan"], row["tabungan"],
                row["total_tagihan"], row["nilai_bayar"], row["selisih"],
                row["status"], row["ket1"], row["ket2"]]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(r, j, v)
            cell.font = _BASE; cell.border = _BORDER
            if j in _MONEY_COLS:
                cell.number_format = '#,##0'; cell.alignment = _R
            elif j == 4:
                cell.number_format = 'dd/mm/yyyy'; cell.alignment = _C
            elif j == 5:
                cell.number_format = 'hh:mm:ss'; cell.alignment = _C
            elif j in (1, 2, 6, 13):
                cell.alignment = _C
            else:
                cell.alignment = _L
        if row["status"] == STATUS_SESUAI:
            ws.cell(r, 13).fill = _OK
    last = r0 + len(shown) - 1

    gt = last + 1
    ws.cell(gt, 3, "TOTAL").alignment = _L
    ws.cell(gt, 3).font = _BOLD
    for j in _MONEY_COLS:
        col = get_column_letter(j)
        cell = ws.cell(gt, j, f"=SUM({col}{r0}:{col}{last})" if shown else 0)
        cell.number_format = '#,##0'; cell.alignment = _R
    for j in range(1, len(_HEADERS) + 1):
        cell = ws.cell(gt, j); cell.fill = _TFILL; cell.border = _BORDER; cell.font = _BOLD

    widths = [5, 12, 28, 13, 10, 8, 12, 12, 12, 14, 14, 12, 10, 16, 16]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A5"

    _build_ringkasan(wb, rows, shown, meta, only_sesuai)
    return wb


def _build_ringkasan(wb, rows, shown, meta, only_sesuai):
    rs = wb.create_sheet("Ringkasan", 0)
    rs.sheet_view.showGridLines = False
    rs["A1"] = "RINGKASAN REKONSILIASI"; rs["A1"].font = _TITLE

    n_sesuai = sum(1 for r in rows if r["status"] == STATUS_SESUAI)
    n_kurang = sum(1 for r in rows if r["status"] == STATUS_KURANG)
    n_lebih = sum(1 for r in rows if r["status"] == STATUS_LEBIH)
    n_unmatched = sum(1 for r in rows if not r["matched"])
    total_bayar_shown = sum(r["nilai_bayar"] for r in shown)

    info = [
        ("Laporan", "R-5401 — Transaksi via E-Banking & Counter"),
        ("Nama Perusahaan", f"{meta.get('kode','') or '—'}-{meta.get('nama_pt','') or ''}".strip("-")),
        ("Tanggal", meta.get("tanggal_label", "") or "—"),
        ("Total Transaksi (laporan)", len(rows)),
        (f"Ditampilkan di sheet Transaksi{' (hanya Sesuai)' if only_sesuai else ''}", len(shown)),
        ("— Rincian Status (semua transaksi laporan) —", ""),
        (STATUS_SESUAI, n_sesuai),
        (STATUS_KURANG, n_kurang),
        (STATUS_LEBIH, n_lebih),
        ("Tanpa data master (No. Pelanggan tak ditemukan)", n_unmatched),
        ("Total Nilai Bayar (yang ditampilkan)", total_bayar_shown),
    ]
    r = 3
    for k, v in info:
        a = rs.cell(r, 1, k); b = rs.cell(r, 2, v)
        head = str(k).startswith("—")
        a.font = Font(_A, size=10, bold=head); b.font = _BOLD if isinstance(v, int) else _BASE
        fill = _TFILL if head else _OK
        a.fill = fill; b.fill = fill; a.border = _BORDER; b.border = _BORDER
        a.alignment = _L
        if isinstance(v, int):
            b.number_format = '#,##0'; b.alignment = _R
        else:
            b.alignment = _L
        r += 1
    rs.column_dimensions["A"].width = 40
    rs.column_dimensions["B"].width = 28
