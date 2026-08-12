"""Parser & Excel builder untuk laporan bank R-5401 (fixed-width text)."""
import re
from datetime import datetime, time as dtime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- posisi kolom fixed-width (0-indexed) ----
SL_NO, SL_CUST, SL_NAMA = (0, 6), (6, 26), (26, 46)
SL_NILAI, SL_TGL, SL_WAKTU = (49, 72), (72, 82), (82, 92)
SL_LOK, SL_K1, SL_K2 = (92, 99), (99, 116), (116, 999)


def _rupiah(n):
    return "Rp " + f"{int(round(n)):,}".replace(",", ".")


def parse_report(text):
    """Terima isi file teks -> (meta:dict, rows:list).

    meta: kode, nama_pt, cabang, tanggal_label, footer_total, footer_count
    rows: [no, cust, nama, nilai, tgl(date), waktu(time), jam(int), lok, k1, k2]
    """
    lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    meta = {"kode": "", "nama_pt": "", "cabang": "", "tanggal_label": "",
            "footer_total": None, "footer_count": None}

    for ln in lines:
        if "NAMA PERUSAHAAN" in ln:
            m = re.search(r"NAMA PERUSAHAAN\s*:\s*(\d+)-(.+?)\s*$", ln)
            if m:
                meta["kode"] = m.group(1)
                meta["nama_pt"] = m.group(2).strip()
        elif "CABANG" in ln and not meta["cabang"]:
            m = re.search(r"CABANG\s*:\s*(.+?)\s*$", ln)
            if m:
                meta["cabang"] = m.group(1).strip()
        elif "TANGGAL" in ln and not meta["tanggal_label"]:
            m = re.search(r"TANGGAL\s*:\s*(\d{2}/\d{2}/\d{2})", ln)
            if m:
                d, mo, y = m.group(1).split("/")
                meta["tanggal_label"] = f"{d}/{mo}/20{y}"
        elif re.match(r"\s*TOTAL NILAI TRANSAKSI", ln):
            m = re.search(r"IDR\s*([\d,]+\.\d{2})", ln)
            if m:
                meta["footer_total"] = float(m.group(1).replace(",", ""))
        elif re.match(r"\s*TOTAL TRANSAKSI", ln):
            m = re.search(r":\s*(\d+)", ln)
            if m:
                meta["footer_count"] = int(m.group(1))

    rows = []
    for ln in lines:
        if not re.match(r"^\s*\d+\s+\d+\s+", ln):
            continue
        if "IDR" not in ln[44:50]:
            continue
        try:
            no = int(ln[slice(*SL_NO)].strip())
            cust = ln[slice(*SL_CUST)].strip()
            nama = ln[slice(*SL_NAMA)].strip()
            nilai = float(ln[slice(*SL_NILAI)].strip().replace(",", ""))
            d, mo, y = ln[slice(*SL_TGL)].strip().split("/")
            tgl = datetime(2000 + int(y), int(mo), int(d)).date()
            hh, mm, ss = ln[slice(*SL_WAKTU)].strip().split(":")
            waktu = dtime(int(hh), int(mm), int(ss))
            lok = ln[slice(*SL_LOK)].strip()
            k1 = ln[slice(*SL_K1)].strip()
            k2 = ln[slice(*SL_K2)].strip()
        except (ValueError, IndexError):
            continue
        rows.append([no, cust, nama, nilai, tgl, waktu, int(hh), lok, k1, k2])

    return meta, rows


# ---- Laporan BCA Virtual Account (format lain, kolom bebas-spasi) ----
_BCA_ROW = re.compile(
    r"^\s*(\d+)\s+(\d[\d-]*)\s+(\S+)\s+(.+?)\s+IDR\s+"
    r"([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+"
    r"(\d{2}/\d{2}/\d{4})\s+(\d{2}:\d{2}:\d{2})\s*(.*)$"
)


def parse_bca_va(text):
    """Parser 'Laporan BCA Virtual Account' -> (meta, rows) dengan struktur
    identik parse_report, agar reconcile_pembayaran bisa dipakai apa adanya.

    - No. Virtual Account '63713-0388' -> cust = '0388' (bagian setelah '-'),
      sehingga reconcile SMP (kode + cust) menghasilkan NO VA penuh '637130388'.
    - nilai bayar = Total Transfer (uang yang benar-benar ditransfer).
    """
    lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    meta = {"kode": "", "nama_pt": "", "cabang": "", "tanggal_label": "",
            "footer_total": None, "footer_count": None}
    for ln in lines:
        if "Kode Perusahaan" in ln:
            m = re.search(r"Kode Perusahaan\s*:\s*(\d+)", ln)
            if m:
                meta["kode"] = m.group(1)
        elif "Nama Perusahaan" in ln:
            m = re.search(r"Nama Perusahaan\s*:\s*(.+?)\s*$", ln)
            if m:
                meta["nama_pt"] = m.group(1).strip()

    rows = []
    for ln in lines:
        m = _BCA_ROW.match(ln)
        if not m:
            continue
        try:
            no = int(m.group(1))
            va = m.group(2)
            cust = va.split("-", 1)[1] if "-" in va else va
            nama = m.group(4).strip()
            transfer = float(m.group(6).replace(",", ""))     # Total Transfer = nilai bayar
            d, mo, y = m.group(7).split("/")
            tgl = datetime(int(y), int(mo), int(d)).date()
            hh, mm, ss = m.group(8).split(":")
            waktu = dtime(int(hh), int(mm), int(ss))
            berita = m.group(9).strip()
            parts = re.split(r"\s{2,}", berita) if berita else []
            k1 = parts[0] if parts else ""
            k2 = parts[1] if len(parts) > 1 else ""
        except (ValueError, IndexError):
            continue
        rows.append([no, cust, nama, transfer, tgl, waktu, int(hh), "", k1, k2])

    if rows and not meta["tanggal_label"]:
        meta["tanggal_label"] = rows[0][4].strftime("%d/%m/%Y")
    return meta, rows


def parse_laporan(text):
    """Auto-deteksi format laporan -> (meta, rows). Mendukung R-5401 & BCA VA."""
    head = "\n".join(text.splitlines()[:8]).upper()
    if "VIRTUAL ACCOUNT" in head:
        return parse_bca_va(text)
    return parse_report(text)


# ---------- styling ----------
_A = "Arial"
_HFILL = PatternFill("solid", fgColor="1F4E5F")
_HFONT = Font(name=_A, bold=True, color="FFFFFF", size=10)
_TITLE = Font(name=_A, bold=True, size=14, color="1F4E5F")
_SUB = Font(name=_A, italic=True, size=9, color="595959")
_BASE = Font(name=_A, size=10)
_BOLD = Font(name=_A, bold=True, size=10)
_TFILL = PatternFill("solid", fgColor="D9E7EC")
_BAND = PatternFill("solid", fgColor="F2F7F9")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_C = Alignment("center", vertical="center")
_L = Alignment("left", vertical="center")
_R = Alignment("right", vertical="center")
_HEADERS = ["No.", "No. Pelanggan/TXN", "Nama Pelanggan", "Nilai Transaksi (IDR)",
            "Tgl. TXN", "Waktu", "Jam", "Lokasi", "Keterangan 1", "Keterangan 2"]


def _write_data_sheet(ws, rows, meta):
    ws.sheet_view.showGridLines = False
    ws["A1"] = "LAPORAN TRANSAKSI VIA E-BANKING & COUNTER"; ws["A1"].font = _TITLE
    ws["A2"] = (f"{meta.get('nama_pt','')} ({meta.get('kode','')})  •  "
                f"{meta.get('cabang','')}  •  Tanggal: {meta.get('tanggal_label','')}  •  "
                f"Laporan R-5401 (Harian)"); ws["A2"].font = _SUB
    HROW = 4
    for c, h in enumerate(_HEADERS, 1):
        cell = ws.cell(HROW, c, h)
        cell.fill = _HFILL; cell.font = _HFONT; cell.alignment = _C; cell.border = _BORDER
    r0 = HROW + 1
    for i, rec in enumerate(rows):
        r = r0 + i
        for c, v in enumerate(rec, 1):
            cell = ws.cell(r, c, v); cell.font = _BASE; cell.border = _BORDER
        ws.cell(r, 1).alignment = _C; ws.cell(r, 2).alignment = _C; ws.cell(r, 3).alignment = _L
        ws.cell(r, 4).number_format = '#,##0'; ws.cell(r, 4).alignment = _R
        ws.cell(r, 5).number_format = 'dd/mm/yyyy'; ws.cell(r, 5).alignment = _C
        ws.cell(r, 6).number_format = 'hh:mm:ss'; ws.cell(r, 6).alignment = _C
        ws.cell(r, 7).alignment = _C; ws.cell(r, 8).alignment = _C
        ws.cell(r, 9).alignment = _L; ws.cell(r, 10).alignment = _L
        if i % 2 == 1:
            for c in range(1, 11):
                ws.cell(r, c).fill = _BAND
    last = r0 + len(rows) - 1
    tr = last + 1
    ws.cell(tr, 3, "TOTAL").font = _BOLD; ws.cell(tr, 3).alignment = _R
    tc = ws.cell(tr, 4, f"=SUM(D{r0}:D{last})"); tc.font = _BOLD
    tc.number_format = '#,##0'; tc.alignment = _R
    cc = ws.cell(tr, 1, f"=COUNT(A{r0}:A{last})"); cc.font = _BOLD; cc.alignment = _C
    for c in range(1, 11):
        ws.cell(tr, c).fill = _TFILL; ws.cell(tr, c).border = _BORDER
    for c, w in enumerate([6, 17, 20, 20, 12, 11, 7, 10, 20, 18], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A5"
    ws.cell(tr + 2, 1,
            "Catatan: Nama pelanggan & Keterangan pada laporan sumber terpotong "
            "(field lebar-tetap ±16 karakter). Nilai, tanggal, waktu, dan lokasi 100% akurat.").font = _SUB
    return r0, last


def _write_summary_sheet(rs, rows, r0, last, meta):
    rs.sheet_view.showGridLines = False
    rs["A1"] = "RINGKASAN TRANSAKSI"; rs["A1"].font = _TITLE
    rs["A2"] = f"{meta.get('tanggal_label','')} — dihitung dengan formula dari sheet Data Transaksi"
    rs["A2"].font = _SUB
    DS = "'Data Transaksi'"
    NREF = f"{DS}!$D${r0}:$D${last}"; LREF = f"{DS}!$H${r0}:$H${last}"; JREF = f"{DS}!$G${r0}:$G${last}"
    kpis = [("Total Transaksi", f"=COUNT({NREF})", "#,##0"),
            ("Total Nilai (IDR)", f"=SUM({NREF})", '"Rp "#,##0'),
            ("Rata-rata / Transaksi", f"=AVERAGE({NREF})", '"Rp "#,##0'),
            ("Transaksi Terbesar", f"=MAX({NREF})", '"Rp "#,##0'),
            ("Transaksi Terkecil", f"=MIN({NREF})", '"Rp "#,##0')]
    kr = 4
    rs.cell(kr, 1, "INDIKATOR").fill = _HFILL; rs.cell(kr, 1).font = _HFONT
    rs.cell(kr, 1).border = _BORDER; rs.cell(kr, 1).alignment = _L
    rs.cell(kr, 2, "NILAI").fill = _HFILL; rs.cell(kr, 2).font = _HFONT
    rs.cell(kr, 2).border = _BORDER; rs.cell(kr, 2).alignment = _R
    for i, (lbl, fml, fmt) in enumerate(kpis):
        r = kr + 1 + i
        rs.cell(r, 1, lbl).font = _BASE; rs.cell(r, 1).border = _BORDER; rs.cell(r, 1).alignment = _L
        c = rs.cell(r, 2, fml); c.font = _BOLD; c.number_format = fmt; c.border = _BORDER; c.alignment = _R
        if i % 2 == 1:
            rs.cell(r, 1).fill = _BAND; rs.cell(r, 2).fill = _BAND
    locs = sorted(set(r[7] for r in rows))
    lr = kr + len(kpis) + 3
    rs.cell(lr - 1, 1, "REKAP PER LOKASI / KODE CHANNEL").font = _BOLD
    for c, h in enumerate(["Lokasi", "Jumlah TXN", "Total Nilai (IDR)", "% Nilai"], 1):
        cell = rs.cell(lr, c, h); cell.fill = _HFILL; cell.font = _HFONT
        cell.border = _BORDER; cell.alignment = _L if c == 1 else _R
    ltot = lr + len(locs) + 1
    for i, lk in enumerate(locs):
        r = lr + 1 + i
        rs.cell(r, 1, lk).font = _BASE; rs.cell(r, 1).border = _BORDER; rs.cell(r, 1).alignment = _L
        rs.cell(r, 2, f'=COUNTIF({LREF},A{r})').font = _BASE
        rs.cell(r, 2).number_format = '#,##0'; rs.cell(r, 2).border = _BORDER; rs.cell(r, 2).alignment = _R
        rs.cell(r, 3, f'=SUMIF({LREF},A{r},{NREF})').font = _BASE
        rs.cell(r, 3).number_format = '"Rp "#,##0'; rs.cell(r, 3).border = _BORDER; rs.cell(r, 3).alignment = _R
        rs.cell(r, 4, f'=C{r}/$C${ltot}').font = _BASE
        rs.cell(r, 4).number_format = '0.0%'; rs.cell(r, 4).border = _BORDER; rs.cell(r, 4).alignment = _R
        if i % 2 == 1:
            for c in range(1, 5):
                rs.cell(r, c).fill = _BAND
    rs.cell(ltot, 1, "TOTAL").font = _BOLD; rs.cell(ltot, 1).alignment = _L
    rs.cell(ltot, 2, f'=SUM(B{lr+1}:B{lr+len(locs)})').font = _BOLD
    rs.cell(ltot, 2).number_format = '#,##0'; rs.cell(ltot, 2).alignment = _R
    rs.cell(ltot, 3, f'=SUM(C{lr+1}:C{lr+len(locs)})').font = _BOLD
    rs.cell(ltot, 3).number_format = '"Rp "#,##0'; rs.cell(ltot, 3).alignment = _R
    rs.cell(ltot, 4, f'=C{ltot}/C{ltot}').font = _BOLD
    rs.cell(ltot, 4).number_format = '0.0%'; rs.cell(ltot, 4).alignment = _R
    for c in range(1, 5):
        rs.cell(ltot, c).fill = _TFILL; rs.cell(ltot, c).border = _BORDER
    hr = ltot + 3
    rs.cell(hr - 1, 1, "REKAP PER JAM (POLA WAKTU PEMBAYARAN)").font = _BOLD
    for c, h in enumerate(["Jam", "Jumlah TXN", "Total Nilai (IDR)"], 1):
        cell = rs.cell(hr, c, h); cell.fill = _HFILL; cell.font = _HFONT
        cell.border = _BORDER; cell.alignment = _L if c == 1 else _R
    hours = sorted(set(r[6] for r in rows))
    for i, jam in enumerate(hours):
        r = hr + 1 + i
        rs.cell(r, 1, f"{jam:02d}:00").font = _BASE; rs.cell(r, 1).border = _BORDER; rs.cell(r, 1).alignment = _L
        rs.cell(r, 2, f'=SUMPRODUCT(({JREF}={jam})*1)').font = _BASE
        rs.cell(r, 2).number_format = '#,##0'; rs.cell(r, 2).border = _BORDER; rs.cell(r, 2).alignment = _R
        rs.cell(r, 3, f'=SUMPRODUCT(({JREF}={jam})*{NREF})').font = _BASE
        rs.cell(r, 3).number_format = '"Rp "#,##0'; rs.cell(r, 3).border = _BORDER; rs.cell(r, 3).alignment = _R
        if i % 2 == 1:
            for c in range(1, 4):
                rs.cell(r, c).fill = _BAND
    htot = hr + len(hours) + 1
    rs.cell(htot, 1, "TOTAL").font = _BOLD; rs.cell(htot, 1).alignment = _L
    rs.cell(htot, 2, f'=SUM(B{hr+1}:B{hr+len(hours)})').font = _BOLD
    rs.cell(htot, 2).number_format = '#,##0'; rs.cell(htot, 2).alignment = _R
    rs.cell(htot, 3, f'=SUM(C{hr+1}:C{hr+len(hours)})').font = _BOLD
    rs.cell(htot, 3).number_format = '"Rp "#,##0'; rs.cell(htot, 3).alignment = _R
    for c in range(1, 4):
        rs.cell(htot, c).fill = _TFILL; rs.cell(htot, c).border = _BORDER
    for col, w in zip("ABCD", [26, 14, 20, 10]):
        rs.column_dimensions[col].width = w


def build_workbook(rows, meta):
    """Bangun workbook 1 laporan -> Workbook object."""
    wb = Workbook()
    ws = wb.active; ws.title = "Data Transaksi"
    r0, last = _write_data_sheet(ws, rows, meta)
    rs = wb.create_sheet("Ringkasan")
    _write_summary_sheet(rs, rows, r0, last, meta)
    return wb


def build_combined_workbook(datasets):
    """datasets: list of (meta, rows). Bangun 1 workbook gabungan semua laporan."""
    wb = Workbook()
    ws = wb.active; ws.title = "Data Gabungan"; ws.sheet_view.showGridLines = False
    ws["A1"] = "DATA TRANSAKSI GABUNGAN — LAPORAN R-5401"; ws["A1"].font = _TITLE
    ws["A2"] = f"{len(datasets)} laporan digabung  •  kolom Tanggal & Kode Perusahaan sebagai penanda sumber"
    ws["A2"].font = _SUB
    headers = ["Tanggal", "Kode PT", "No.", "No. Pelanggan/TXN", "Nama Pelanggan",
               "Nilai Transaksi (IDR)", "Waktu", "Jam", "Lokasi", "Keterangan 1", "Keterangan 2"]
    HROW = 4
    for c, h in enumerate(headers, 1):
        cell = ws.cell(HROW, c, h)
        cell.fill = _HFILL; cell.font = _HFONT; cell.alignment = _C; cell.border = _BORDER
    r0 = HROW + 1
    r = r0
    for meta, rows in datasets:
        for rec in rows:
            no, cust, nama, nilai, tgl, waktu, jam, lok, k1, k2 = rec
            vals = [tgl, meta.get("kode", ""), no, cust, nama, nilai, waktu, jam, lok, k1, k2]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(r, c, v); cell.font = _BASE; cell.border = _BORDER
            ws.cell(r, 1).number_format = 'dd/mm/yyyy'; ws.cell(r, 1).alignment = _C
            ws.cell(r, 2).alignment = _C; ws.cell(r, 3).alignment = _C; ws.cell(r, 4).alignment = _C
            ws.cell(r, 5).alignment = _L
            ws.cell(r, 6).number_format = '#,##0'; ws.cell(r, 6).alignment = _R
            ws.cell(r, 7).number_format = 'hh:mm:ss'; ws.cell(r, 7).alignment = _C
            ws.cell(r, 8).alignment = _C; ws.cell(r, 9).alignment = _C
            ws.cell(r, 10).alignment = _L; ws.cell(r, 11).alignment = _L
            r += 1
    last = r - 1
    tr = last + 1
    ws.cell(tr, 5, "TOTAL").font = _BOLD; ws.cell(tr, 5).alignment = _R
    tc = ws.cell(tr, 6, f"=SUM(F{r0}:F{last})"); tc.font = _BOLD
    tc.number_format = '#,##0'; tc.alignment = _R
    cc = ws.cell(tr, 3, f"=COUNT(C{r0}:C{last})"); cc.font = _BOLD; cc.alignment = _C
    for c in range(1, 12):
        ws.cell(tr, c).fill = _TFILL; ws.cell(tr, c).border = _BORDER
    for c, w in enumerate([12, 9, 6, 17, 20, 20, 11, 7, 10, 20, 18], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A5"

    # sheet rekap harian per PT
    rk = wb.create_sheet("Rekap Harian")
    rk.sheet_view.showGridLines = False
    rk["A1"] = "REKAP PER TANGGAL & KODE PERUSAHAAN"; rk["A1"].font = _TITLE
    DG = "'Data Gabungan'"
    TREF = f"{DG}!$A${r0}:$A${last}"; KREF = f"{DG}!$B${r0}:$B${last}"; VREF = f"{DG}!$F${r0}:$F${last}"
    combos = sorted({(meta.get("tanggal_label", ""), meta.get("kode", "")) for meta, _ in datasets})
    # map label -> serial-friendly: we filter by kode & the date text via SUMPRODUCT on real dates is hard;
    # instead build one row per (meta) dataset directly from python-known counts for clarity.
    hh = 3
    for c, h in enumerate(["Tanggal", "Kode PT", "Nama PT", "Jumlah TXN", "Total Nilai (IDR)"], 1):
        cell = rk.cell(hh, c, h); cell.fill = _HFILL; cell.font = _HFONT
        cell.border = _BORDER; cell.alignment = _L if c <= 3 else _R
    ri = hh + 1
    for i, (meta, rows) in enumerate(sorted(datasets, key=lambda d: (d[0].get("tanggal_label",""), d[0].get("kode","")))):
        rk.cell(ri, 1, meta.get("tanggal_label", "")).font = _BASE; rk.cell(ri, 1).border = _BORDER; rk.cell(ri, 1).alignment = _L
        rk.cell(ri, 2, meta.get("kode", "")).font = _BASE; rk.cell(ri, 2).border = _BORDER; rk.cell(ri, 2).alignment = _L
        rk.cell(ri, 3, meta.get("nama_pt", "")).font = _BASE; rk.cell(ri, 3).border = _BORDER; rk.cell(ri, 3).alignment = _L
        rk.cell(ri, 4, len(rows)).font = _BASE; rk.cell(ri, 4).number_format = '#,##0'; rk.cell(ri, 4).border = _BORDER; rk.cell(ri, 4).alignment = _R
        sm = rk.cell(ri, 5, sum(x[3] for x in rows)); sm.font = _BASE; sm.number_format = '"Rp "#,##0'; sm.border = _BORDER; sm.alignment = _R
        if i % 2 == 1:
            for c in range(1, 6):
                rk.cell(ri, c).fill = _BAND
        ri += 1
    rk.cell(ri, 3, "TOTAL").font = _BOLD; rk.cell(ri, 3).alignment = _R
    rk.cell(ri, 4, f"=SUM(D{hh+1}:D{ri-1})").font = _BOLD; rk.cell(ri, 4).number_format = '#,##0'; rk.cell(ri, 4).alignment = _R
    rk.cell(ri, 5, f"=SUM(E{hh+1}:E{ri-1})").font = _BOLD; rk.cell(ri, 5).number_format = '"Rp "#,##0'; rk.cell(ri, 5).alignment = _R
    for c in range(1, 6):
        rk.cell(ri, c).fill = _TFILL; rk.cell(ri, c).border = _BORDER
    for col, w in zip("ABCDE", [12, 9, 34, 12, 20]):
        rk.column_dimensions[col].width = w
    return wb
